# utils/pipeline_logger.py
# PURPOSE: Record pipeline failures and flag emails for human review.
#
# TWO OUTPUTS:
#   data/failed_emails.log    — plain-text log, one line per failure.
#                               Open in any text editor to see what went wrong.
#   data/review_queue.xlsx    — spreadsheet of emails needing manual attention:
#                               - AI extraction completely failed
#                               - AI confidence below CONFIDENCE_REVIEW_THRESHOLD
#                               - AI flagged important fields as missing
#
# DESIGN RULE:
#   These functions never raise exceptions themselves. A logging failure must
#   never crash the main pipeline — it prints a warning and moves on.

import sys
import datetime
import traceback
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from config.fields import CRM_COLUMNS, CONFIDENCE_REVIEW_THRESHOLD

# Output file paths
FAILED_LOG_PATH    = PROJECT_ROOT / "data" / "failed_emails.log"
REVIEW_QUEUE_PATH  = PROJECT_ROOT / "data" / "review_queue.xlsx"

# Review queue has all 55 CRM columns plus two extra audit columns
REVIEW_EXTRA_COLS  = ["queued_at", "failure_reason"]
REVIEW_COLS        = CRM_COLUMNS + REVIEW_EXTRA_COLS

# Visual styling: orange header to distinguish from the blue main CRM file
REVIEW_HEADER_FILL = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
REVIEW_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)


# ---------------------------------------------------------------------------
# LOG FAILURE — append one line to failed_emails.log
# ---------------------------------------------------------------------------

def log_failure(email_record: dict, reason: str) -> None:
    """
    Append a single failure entry to data/failed_emails.log.

    This is a plain-text audit trail. Each line records:
        timestamp | FAILED | email_id | sender | subject (truncated) | reason

    Parameters:
        email_record — raw email dict from email_reader (or any dict with email fields)
        reason       — short description of what went wrong
    """
    try:
        FAILED_LOG_PATH.parent.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        email_id  = email_record.get("email_id", "unknown")[:30]
        sender    = email_record.get("sender_email", "unknown")
        subject   = email_record.get("subject", "(no subject)")[:60]

        # Each log line is pipe-separated for easy parsing later if needed
        log_line = (
            f"[{timestamp}] FAILED"
            f" | id={email_id}"
            f" | from={sender}"
            f" | subject={subject}"
            f" | reason={reason}\n"
        )

        # 'a' mode = append — never overwrites existing log entries
        with open(FAILED_LOG_PATH, "a", encoding="utf-8") as f:
            f.write(log_line)

    except Exception as e:
        # Logging must never crash the pipeline
        print(f"  [pipeline_logger] WARNING: Could not write to failed_emails.log — {e}")


# ---------------------------------------------------------------------------
# ADD TO REVIEW QUEUE — append a row to review_queue.xlsx
# ---------------------------------------------------------------------------

def add_to_review_queue(row_dict: dict, reason: str) -> None:
    """
    Append one email record to data/review_queue.xlsx for human review.

    Called when:
        - AI extraction completely failed (no output at all)
        - AI confidence < CONFIDENCE_REVIEW_THRESHOLD (currently {threshold})
        - AI flagged missing required fields
        - An unexpected exception occurred during processing

    The row_dict can be a full 55-field CRM row OR just a partial email record —
    missing keys are filled with empty strings automatically.

    Deduplicates on email_id so re-running the pipeline won't create duplicate
    review entries for the same email.

    Parameters:
        row_dict — dict with any subset of CRM fields + raw email metadata
        reason   — human-readable explanation: "low confidence (0.42) | missing: deadline"
    """.format(threshold=CONFIDENCE_REVIEW_THRESHOLD)

    try:
        REVIEW_QUEUE_PATH.parent.mkdir(parents=True, exist_ok=True)

        # Build a complete review row — fill any missing CRM columns with ""
        review_row = {}
        for col in CRM_COLUMNS:
            review_row[col] = row_dict.get(col, "")

        # Add the two extra audit columns
        review_row["queued_at"]     = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        review_row["failure_reason"] = reason

        new_df = pd.DataFrame([review_row])[REVIEW_COLS]

        if REVIEW_QUEUE_PATH.exists():
            existing_df = pd.read_excel(REVIEW_QUEUE_PATH, engine="openpyxl", dtype=str)
            existing_df = existing_df.fillna("")

            # Ensure all columns exist in the existing file (handles schema additions)
            for col in REVIEW_COLS:
                if col not in existing_df.columns:
                    existing_df[col] = ""
            existing_df = existing_df[REVIEW_COLS]

            # Deduplicate: skip if this email_id is already in the review queue
            existing_ids = set(existing_df["email_id"].tolist())
            email_id     = str(review_row.get("email_id", ""))
            if email_id and email_id in existing_ids:
                return  # Already in queue — no duplicate

            combined_df = pd.concat([existing_df, new_df], ignore_index=True)
        else:
            combined_df = new_df

        combined_df.to_excel(REVIEW_QUEUE_PATH, index=False, engine="openpyxl")
        _format_review_queue(REVIEW_QUEUE_PATH)

    except Exception as e:
        print(f"  [pipeline_logger] WARNING: Could not write to review_queue.xlsx — {e}")
        print(f"  [pipeline_logger] Traceback: {traceback.format_exc()}")


# ---------------------------------------------------------------------------
# FORMATTING
# ---------------------------------------------------------------------------

def _format_review_queue(filepath: Path) -> None:
    """
    Apply formatting to the review queue Excel file.
    Orange header (instead of the CRM's blue) makes it visually distinct.
    Also adds a note in cell A1 area via header color.
    """
    try:
        wb = load_workbook(filepath)
        ws = wb.active

        # Orange header row to signal "this needs attention"
        for col_idx in range(1, len(REVIEW_COLS) + 1):
            cell           = ws.cell(row=1, column=col_idx)
            cell.font      = REVIEW_HEADER_FONT
            cell.fill      = REVIEW_HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Column widths — wider for key fields, standard for the rest
        key_widths = {
            "failure_reason": 50,
            "queued_at":      20,
            "email_summary":  50,
            "next_action":    40,
            "subject":        40,
            "sender_email":   28,
            "client_name":    22,
            "designation":    28,
        }
        for col_idx, col_name in enumerate(REVIEW_COLS, start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = key_widths.get(col_name, 14)

        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 20

        # Highlight failure_reason column in data rows so it stands out
        fr_idx = REVIEW_COLS.index("failure_reason") + 1
        ALERT_FILL = PatternFill(start_color="FFE5B3", end_color="FFE5B3", fill_type="solid")
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=fr_idx).fill = ALERT_FILL
            ws.row_dimensions[row_idx].height = 30

        wb.save(filepath)

    except Exception as e:
        # Formatting failure is non-fatal — the data is already saved
        print(f"  [pipeline_logger] WARNING: Could not format review_queue.xlsx — {e}")


# ---------------------------------------------------------------------------
# SUMMARY HELPER — used by run_parser to print end-of-run stats
# ---------------------------------------------------------------------------

def get_review_queue_count() -> int:
    """Return the number of records currently in the review queue. Returns 0 on error."""
    try:
        if not REVIEW_QUEUE_PATH.exists():
            return 0
        df = pd.read_excel(REVIEW_QUEUE_PATH, engine="openpyxl")
        return len(df)
    except Exception:
        return 0


# ---------------------------------------------------------------------------
# QUICK TEST
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    print("=== pipeline_logger.py self-test ===\n")

    test_email = {
        "email_id":     "test-logger-001",
        "received_date":"2026-06-04",
        "subject":      "Manpower Request - Drilling Engineers",
        "sender_email": "ahmed@aramco.com",
        "sender_name":  "Ahmed Al-Rashid",
        "client_name":  "Saudi Aramco",
        "designation":  "Drilling Engineer",
        "llm_confidence": 0.45,
        "missing_fields_flag": "deadline, rates",
        "status":       "New",
        "review_status":"Pending Review",
    }

    print("  Writing to failed_emails.log...")
    log_failure(test_email, reason="Groq API timeout after 3 retries")
    print(f"  Log path: {FAILED_LOG_PATH}")

    print("  Writing to review_queue.xlsx...")
    add_to_review_queue(test_email, reason="low confidence (0.45) | missing: deadline, rates")
    print(f"  Queue path: {REVIEW_QUEUE_PATH}")
    print(f"  Queue count: {get_review_queue_count()}")

    # Verify deduplication — second call with same email_id should not add a row
    count_before = get_review_queue_count()
    add_to_review_queue(test_email, reason="duplicate test")
    count_after = get_review_queue_count()
    assert count_after == count_before, "Deduplication failed!"
    print("  PASS: deduplication works (same email_id not added twice)")

    # Check the log file has content
    log_content = FAILED_LOG_PATH.read_text(encoding="utf-8")
    assert "test-logger-001" in log_content
    assert "Groq API timeout" in log_content
    print("  PASS: log file contains expected content")

    print("\n  Self-test complete. Check data/ folder for output files.")
