# storage.py - saves CRM records to Excel (primary) and Supabase (cloud backup)
# Phase 14: Supabase added as second destination

import os
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from dotenv import load_dotenv
from supabase import create_client

load_dotenv()

# --- File paths ---
DATA_DIR = Path(__file__).parent.parent / "data"
EXCEL_PATH = DATA_DIR / "crm_data.xlsx"
CRM_FILE_PATH = EXCEL_PATH  # alias for backwards compatibility

# --- Supabase client setup ---
SUPABASE_URL = os.getenv("SUPABASE_URL")
# Service role key bypasses Row Level Security — correct for server-side scripts.
# Anon key is for browser clients and is blocked by RLS on INSERT.
# Add SUPABASE_SERVICE_KEY to .env to enable full write access.
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_KEY") or os.getenv("SUPABASE_KEY")

# Supabase column names that can be updated, and any remapping from internal names
_SUPABASE_COLS = {
    "email_id", "received_date", "sender_name", "sender_email", "subject",
    "category", "client", "project", "positions", "headcount", "location",
    "mobilization_date", "duration", "rates", "contact_person", "contact_email",
    "deadline", "summary", "suggested_action", "status", "psl_categories",
    "classification", "confidence_score",
}
# Internal field name -> Supabase column name (when they differ)
_SUPABASE_RENAME = {
    "email_summary": "summary",
    "next_action":   "suggested_action",
}

def _sanitize_for_supabase(updates: dict) -> dict:
    """Keep only Supabase-known columns, renaming any internal names that differ."""
    result = {}
    for k, v in updates.items():
        col = _SUPABASE_RENAME.get(k, k)
        if col in _SUPABASE_COLS:
            result[col] = v
    return result

if SUPABASE_URL and SUPABASE_KEY:
    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
    print("Supabase connected")
else:
    supabase = None
    print("Supabase credentials not found - cloud save disabled")


def save_to_supabase(record: dict):
    if not supabase:
        print("   Supabase not connected - skipping cloud save")
        return

    # confidence_score may be a float, empty string, or None — guard the cast
    try:
        conf = float(record.get("llm_confidence") or 0.0)
    except (ValueError, TypeError):
        conf = 0.0

    supabase_record = {
        # --- Identity — use req_id (REQ-YYYYMMDD-HASH-01) as the stable conflict key.
        # Falls back to raw email_id for any legacy records that pre-date req_id.
        "email_id":          record.get("req_id") or record.get("email_id", ""),
        "received_date":     record.get("received_date", ""),
        "sender_name":       record.get("sender_name", ""),
        "sender_email":      record.get("sender_email", ""),
        "subject":           record.get("subject", ""),
        # --- Client (extractor uses client_name / project_name) ---
        "client":            record.get("client_name", ""),
        "project":           record.get("project_name", ""),
        "contact_person":    record.get("contact_person", ""),
        "contact_email":     record.get("contact_email", ""),
        # --- Role (extractor uses designation, not positions) ---
        "positions":         record.get("designation", ""),
        "headcount":         str(record.get("headcount", "")),
        "psl_categories":    str(record.get("psl_categories", "")),
        # --- Dates / logistics ---
        "location":          record.get("location", ""),
        "mobilization_date": record.get("mobilization_date", ""),
        "duration":          record.get("duration", ""),
        "rates":             record.get("rates", ""),
        "deadline":          record.get("deadline", ""),
        # --- Classification ---
        "category":          record.get("category", ""),
        "classification":    record.get("category", ""),   # same value, different column
        # --- AI output (extractor uses email_summary / next_action) ---
        "summary":           record.get("email_summary", ""),
        "suggested_action":  record.get("next_action", ""),
        "confidence_score":  conf,
        # --- Status ---
        "status":            record.get("status", "New"),
    }

    try:
        supabase.table("crm_requirements").upsert(
            supabase_record,
            on_conflict="email_id"
        ).execute()
        print(f"   Supabase saved: {record.get('subject', 'no subject')[:50]}")
    except Exception as e:
        print(f"   Supabase save failed: {e}")


def save_records(records: list, excel_path: Path = EXCEL_PATH):
    if not records:
        print("No records to save.")
        return

    # Excel first (always) — capture the result so callers can read save counts
    result = save_to_excel(records, excel_path)

    # Supabase second (Phase 14)
    print(f"\nPushing {len(records)} record(s) to Supabase...")
    for record in records:
        save_to_supabase(record)
    print("Supabase sync complete.")

    return result


def update_record(email_id: str, updates: dict, excel_path: Path = EXCEL_PATH):
    """Update an existing record in Excel and Supabase by email_id."""
    # Update Excel
    if excel_path.exists():
        df = pd.read_excel(excel_path, dtype=str)
        mask = df["email_id"].astype(str) == str(email_id)
        if mask.any():
            for key, value in updates.items():
                if key in df.columns:
                    df.loc[mask, key] = str(value)
            df.to_excel(excel_path, index=False)
            _format_excel(excel_path)
            print(f"Excel: record updated -> {email_id}")

    # Update Supabase — use upsert so it inserts the row if it doesn't exist yet,
    # and updates it if it does. .update().eq() silently does nothing on an empty table.
    if supabase:
        try:
            supabase_updates = _sanitize_for_supabase(updates)
            if supabase_updates:
                supabase_updates["email_id"] = email_id  # required for upsert conflict key
                supabase.table("crm_requirements").upsert(
                    supabase_updates, on_conflict="email_id"
                ).execute()
                print(f"   Supabase upserted: {email_id}")
            else:
                print(f"   Supabase: no mapped fields to upsert for {email_id}")
        except Exception as e:
            print(f"   Supabase upsert failed: {e}")

    return {"updated": True}


def save_to_excel(records: list, excel_path: Path = EXCEL_PATH):
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    columns = [
        "email_id", "received_date", "sender_name", "sender_email", "subject",
        "category", "client", "project", "positions", "headcount", "location",
        "mobilization_date", "duration", "rates", "contact_person",
        "contact_email", "deadline", "summary", "suggested_action",
        "status", "psl_categories", "classification", "confidence_score"
    ]

    new_df = pd.DataFrame(records)

    for col in columns:
        if col not in new_df.columns:
            new_df[col] = ""

    new_df = new_df[columns]

    if excel_path.exists():
        existing_df = pd.read_excel(excel_path, dtype=str)
        existing_ids = set(existing_df["email_id"].astype(str))
        truly_new = new_df[~new_df["email_id"].astype(str).isin(existing_ids)]
        combined_df = pd.concat([existing_df, truly_new], ignore_index=True)
        saved_count = len(truly_new)
    else:
        combined_df = new_df
        saved_count = len(new_df)

    combined_df.to_excel(excel_path, index=False)
    _format_excel(excel_path)
    print(f"Excel: {saved_count} new record(s) saved -> {excel_path.name}")

    return {"saved": saved_count, "skipped_duplicates": 0, "total_rows": len(combined_df)}


def _format_excel(excel_path: Path):
    wb = load_workbook(excel_path)
    ws = wb.active

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)

    ws.freeze_panes = "A2"

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    wb.save(excel_path)
