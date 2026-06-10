# run_zavenir_parser.py
#
# PURPOSE: Standalone pipeline runner for the Zavenir Daubert pipeline.
#
# WHAT IT DOES (in order):
#   1. Authenticate with Outlook using utils/auth.py
#   2. Fetch emails from inbox using 03_outlook/email_reader.py
#   3. Filter to only emails where the original sender is tarora@zavenir.com
#      — checks both the sender_email field and the body for forwarded-from patterns
#   4. Download and extract attachment text (PDF / DOCX / XLSX / ZIP) if present
#   5. Clean email body using 04_email_parser/parser.py
#   6. Send to Groq AI via utils/zavenir_extractor.py — extracts commercial fields
#   7. Generate a unique req_id in the same format as the GET Global pipeline
#   8. Save all new records to data/zavenir_crm.xlsx (creates file if not exists)
#   9. Upsert all records to Supabase zavenir_requirements table
#  10. Print a run summary: fetched / found / saved to Excel / saved to Supabase
#
# ISOLATION: This file is completely standalone.
#   - Does NOT import from run_parser.py, storage.py, or any GET Global module
#   - Shares only reusable utilities: auth, email_reader, parser, attachment_extractor
#   - Has its own Excel path, Supabase table, and save logic

import os
import sys
import time
import hashlib
import datetime
import re
import requests
import pandas as pd
from pathlib import Path
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from supabase import create_client

# ---------------------------------------------------------------------------
# PATH SETUP
# All imports below need the project root and sub-folders on sys.path.
# ---------------------------------------------------------------------------

PROJECT_ROOT = Path(__file__).parent
sys.path.insert(0, str(PROJECT_ROOT))
sys.path.insert(0, str(PROJECT_ROOT / "03_outlook"))
sys.path.insert(0, str(PROJECT_ROOT / "04_email_parser"))

# Load .env from project root — same file used by all pipelines
load_dotenv(dotenv_path=PROJECT_ROOT / ".env")

# ---------------------------------------------------------------------------
# IMPORTS — shared utilities (read-only; we never modify these files)
# ---------------------------------------------------------------------------

from utils.auth import get_access_token
from email_reader import fetch_emails, get_email_attachments
import parser as email_parser
from utils.attachment_extractor import extract_text_from_attachment
from utils.zavenir_extractor import (
    extract_zavenir_thread, fetch_conversation, stitch_thread,
    extract_forwarded_sender, get_forwarder_note, detect_assigned_to,
)
from configs.clients.zavenir import FIELDS, EXCEL_OUTPUT, SUPABASE_TABLE, SENDER_FILTER

# ---------------------------------------------------------------------------
# SUPABASE CLIENT — zavenir pipeline uses the same credentials as GET Global
# ---------------------------------------------------------------------------

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_KEY") or os.getenv("SUPABASE_KEY")

if SUPABASE_URL and SUPABASE_KEY:
    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
    print("Supabase connected (zavenir pipeline)")
else:
    supabase = None
    print("WARNING: Supabase credentials not found — cloud save disabled")

# ---------------------------------------------------------------------------
# FILE PATHS
# ---------------------------------------------------------------------------

DATA_DIR   = PROJECT_ROOT / "data"
EXCEL_PATH = PROJECT_ROOT / EXCEL_OUTPUT   # data/zavenir_crm.xlsx

# ---------------------------------------------------------------------------
# SENDER FILTER — pre-compiled for speed inside the per-email loop
# Matches "tarora@zavenir.com" anywhere in the email body text.
# Used to catch emails that Saral forwarded (the sender_email is Saral's
# address, but the original From: header in the body contains Tania's address).
# ---------------------------------------------------------------------------

_ZAVENIR_EMAIL_RE = re.compile(r"tarora@zavenir\.com", re.IGNORECASE)

GRAPH_BASE_URL = "https://graph.microsoft.com/v1.0"

# Set True to print raw Graph API fields and exit without extraction or saves.
DEBUG_DISCOVERY = False
# Set True to run full extraction but skip Excel + Supabase saves.
# Useful for testing prompt or thread changes without touching production data.
DRY_RUN = False


def _search_zavenir_emails(token: str, max_results: int = 25) -> list:
    """
    Search all Outlook folders for emails containing 'tarora@zavenir.com'.

    Uses Graph API $search (full-text) instead of inbox-only fetch, so it finds
    emails regardless of which folder they're in (inbox, archive, subfolders).
    Does NOT apply the domain whitelist — Zavenir emails are handled by this
    pipeline directly, not by email_filters.py.

    Returns a list of dicts in the same format as fetch_emails().
    """
    fields = ",".join([
        "id", "subject", "receivedDateTime", "from", "toRecipients",
        "body", "bodyPreview", "isRead", "hasAttachments", "importance",
        "conversationId", "conversationIndex",   # Phase 17 discovery
    ])
    url = (
        f"{GRAPH_BASE_URL}/me/messages"
        f"?$search=\"tarora@zavenir.com\""
        f"&$select={fields}"
        f"&$top={max_results}"
    )
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }

    try:
        resp = requests.get(url, headers=headers, timeout=30)
        resp.raise_for_status()
    except requests.exceptions.RequestException as e:
        print(f"  [Zavenir search] Graph API error: {e}")
        return []

    raw_emails = resp.json().get("value", [])
    print(f"  [Zavenir search] {len(raw_emails)} email(s) found across all folders.")

    emails = []
    for item in raw_emails:
        sender_info  = item.get("from", {}).get("emailAddress", {})
        to_recipients = [
            {
                "name":    r.get("emailAddress", {}).get("name", ""),
                "address": r.get("emailAddress", {}).get("address", ""),
            }
            for r in item.get("toRecipients", [])
        ]
        emails.append({
            "email_id":        item.get("id", ""),
            "received_date":   item.get("receivedDateTime", ""),
            "subject":         item.get("subject", "(No subject)"),
            "sender_name":     sender_info.get("name", "Unknown"),
            "sender_email":    sender_info.get("address", "Unknown"),
            "to_recipients":   to_recipients,                               # Phase 18
            "body_preview":    item.get("bodyPreview", ""),
            "body_content":    item.get("body", {}).get("content", ""),
            "body_type":       item.get("body", {}).get("contentType", "text"),
            "is_read":           item.get("isRead", False),
            "has_attachments":   item.get("hasAttachments", False),
            "importance":        item.get("importance", "normal"),
            "conversationId":    item.get("conversationId", "NOT PRESENT"),    # Phase 17 discovery
            "conversationIndex": item.get("conversationIndex", "NOT PRESENT"), # Phase 17 discovery
        })
    return emails


# ===========================================================================
# MAIN PIPELINE
# ===========================================================================

def run_pipeline(max_emails: int = 10) -> dict:
    """
    Run the full Zavenir Daubert email pipeline.

    Parameters:
        max_emails — how many emails to fetch from Outlook (keep low for testing)

    Returns:
        A dict with:
            records        — list of extracted CRM row dicts
            emails_fetched — total emails fetched from Outlook
            zavenir_found  — how many of those matched the Zavenir sender filter
    """

    print("=" * 65)
    print("  ZAVENIR DAUBERT PIPELINE")
    print(f"  Sender filter: {SENDER_FILTER}")
    print("  Filter -> Attach -> Clean -> Extract -> Save")
    print("=" * 65)

    # ------------------------------------------------------------------
    # Step 1: Authenticate and search all folders for Zavenir emails
    # Uses Graph API $search so emails in archive/subfolders are found.
    # ------------------------------------------------------------------
    print("\n[Step 1] Searching all Outlook folders for Zavenir emails...")
    token  = get_access_token()
    emails = _search_zavenir_emails(token, max_results=max_emails)

    if not emails:
        print("  No emails found. Check Outlook connection or search index.")
        return {"records": [], "emails_fetched": 0, "zavenir_found": 0}

    print(f"  Found {len(emails)} email(s) matching 'tarora@zavenir.com'.")

    # ------------------------------------------------------------------
    # Phase 17 DISCOVERY — print raw Graph API fields then exit early.
    # No AI extraction, no Excel save, no Supabase upsert.
    # Remove / set DEBUG_DISCOVERY = False when building thread logic.
    # ------------------------------------------------------------------
    if DEBUG_DISCOVERY:

        def _safe(text):
            """Encode to cp1252 so Windows console never throws UnicodeEncodeError."""
            return str(text).encode("cp1252", errors="replace").decode("cp1252")

        print("\n" + "=" * 65)
        print("  [DISCOVERY] STEP 1 — Raw Graph API fields (first 10 emails)")
        print("=" * 65)
        for idx, em in enumerate(emails[:10], start=1):
            print(f"\n  [{idx}]")
            print(f"  email_id         : {_safe(em.get('email_id', '')[:80])}")
            print(f"  subject          : {_safe(em.get('subject', ''))}")
            print(f"  received_date    : {em.get('received_date', '')}")
            print(f"  conversationId   : {_safe(em.get('conversationId', 'NOT PRESENT'))}")
            conv_idx = em.get("conversationIndex", "NOT PRESENT")
            print(f"  conversationIndex: {_safe(conv_idx[:60] if conv_idx != 'NOT PRESENT' else conv_idx)}")

        # STEP 2 — filter for Auto International / Binola / X-Cool / Nox Rust thread
        _THREAD_KW = re.compile(
            r"auto\s*international|binola|x[\s\-]?cool|nox\s*rust",
            re.IGNORECASE,
        )

        def _matches_thread(em: dict) -> bool:
            return any(
                _THREAD_KW.search(em.get(field, "") or "")
                for field in ("subject", "sender_name", "body_preview")
            )

        print("\n" + "=" * 65)
        print("  [DISCOVERY] STEP 2 — Thread keyword filter")
        print("  Keywords: Auto International | Binola | X-Cool | Nox Rust")
        print("=" * 65)
        matches = [em for em in emails if _matches_thread(em)]
        if matches:
            for em in matches:
                print(f"\n  subject          : {_safe(em.get('subject', ''))}")
                print(f"  sender_name      : {_safe(em.get('sender_name', ''))}")
                print(f"  received_date    : {em.get('received_date', '')}")
                print(f"  conversationId   : {_safe(em.get('conversationId', 'NOT PRESENT'))}")
                conv_idx = em.get("conversationIndex", "NOT PRESENT")
                print(f"  conversationIndex: {_safe(conv_idx[:60] if conv_idx != 'NOT PRESENT' else conv_idx)}")
        else:
            print("\n  No keyword matches found in this batch.")

        print("\n" + "=" * 65)
        print("  [DISCOVERY] Complete. No AI extraction. No saves.")
        print("=" * 65 + "\n")
        return {"records": [], "emails_fetched": len(emails), "zavenir_found": 0}

    # ------------------------------------------------------------------
    # Step 2: Filter for Zavenir emails only
    # ------------------------------------------------------------------
    print(f"\n[Step 2] Filtering for sender: {SENDER_FILTER}...")
    zavenir_emails = [e for e in emails if _is_zavenir_email(e)]
    skipped = len(emails) - len(zavenir_emails)

    if skipped:
        print(f"  {skipped} email(s) skipped (not from Zavenir).")
    if not zavenir_emails:
        print("  No Zavenir emails found in this batch.")
        return {"records": [], "emails_fetched": len(emails), "zavenir_found": 0}

    print(f"  {len(zavenir_emails)} Zavenir email(s) found.\n")

    # ------------------------------------------------------------------
    # Step 3: Process each Zavenir email
    # ------------------------------------------------------------------
    records     = []
    failed      = 0

    for i, email in enumerate(zavenir_emails, start=1):
        subject = email.get("subject", "(No subject)")
        sender  = email.get("sender_email", "")
        conv_id = email.get("conversationId", "")

        print(f"[{i}/{len(zavenir_emails)}] {subject[:55]}")
        print(f"           From: {sender}")

        try:
            # ------------------------------------------------------------------
            # Phase 17 — Thread Intelligence
            # Fetch the full conversation using conversationId, then stitch all
            # emails into one text block before passing to the AI.
            # ------------------------------------------------------------------
            if conv_id and conv_id != "NOT PRESENT":
                print("           Fetching conversation thread...")
                thread_emails = fetch_conversation(token, conv_id)
                print(f"           Thread: {len(thread_emails)} email(s) in conversation.")
            else:
                thread_emails = []
                print("           WARNING: No conversationId — using single email body.")

            if thread_emails:
                base_text = stitch_thread(thread_emails)
            else:
                # Fallback: wrap single email in the same [DATE | SENDER] format
                fallback = [{
                    "subject":       email.get("subject", ""),
                    "received_date": email.get("received_date", ""),
                    "sender":        email.get("sender_email", ""),
                    "body_content":  email.get("body_content", ""),
                }]
                base_text = stitch_thread(fallback)

            # --- Attachment text extraction (from the forwarded/anchor email) ---
            if email.get("has_attachments"):
                print("           Attachments: detected — downloading...")
                att_list  = get_email_attachments(token, email["email_id"])
                att_texts = []

                for att in att_list:
                    att_text = extract_text_from_attachment(att["name"], att["bytes"])
                    if att_text:
                        att_texts.append(f"[Attachment: {att['name']}]\n{att_text}")

                if att_texts:
                    combined_att = "\n\n".join(att_texts)
                    base_text   += "\n\n--- ATTACHMENT TEXT ---\n" + combined_att
                    total_chars  = sum(len(t) for t in att_texts)
                    print(
                        f"           Attachments: {len(att_texts)}/{len(att_list)} "
                        f"extracted, {total_chars} chars appended."
                    )
                elif att_list:
                    print(
                        f"           Attachments: {len(att_list)} downloaded but "
                        "no text extracted (scanned image or unsupported format)."
                    )

            # --- AI extraction — multi-product thread mode ---
            print("           Extracting with Groq AI (thread mode)...")
            product_records = extract_zavenir_thread(base_text)

            if not product_records:
                print("           WARNING: AI returned no records — skipping.")
                failed += 1
                continue

            # ------------------------------------------------------------------
            # Phase 18 — Forwarded sender extraction + auto-assignment
            #
            # All Zavenir emails arrive forwarded from tarora@zavenir.com to
            # saral.prabhat@outlook.com. The real customer is named in a
            # "From:" header inside the forwarded block, not in the Graph API
            # `from` field. Use that as sender_name/sender_email instead, and
            # never let SENDER_FILTER (Tania's address) appear as the sender.
            # ------------------------------------------------------------------
            if email.get("body_type") == "html":
                anchor_clean_text = email_parser.strip_html(email.get("body_content", ""))
            else:
                anchor_clean_text = email.get("body_content", "")

            orig_name, orig_email = extract_forwarded_sender(anchor_clean_text)
            if orig_email:
                sender_name  = orig_name or orig_email.split("@")[0]
                sender_email = orig_email
            else:
                sender_name  = email.get("sender_name", "")
                sender_email = email.get("sender_email", "")
                if sender_email.lower() == SENDER_FILTER.lower():
                    sender_name, sender_email = None, None

            forwarder_note = get_forwarder_note(anchor_clean_text)
            assigned_to, assigned_to_confidence = detect_assigned_to(
                forwarder_note, email.get("to_recipients", [])
            )

            # --- Assemble one CRM record per extracted product ---
            # Use conversationId as the hash base so all products from the same
            # thread share a common REQ-…-XXXXXX- prefix and differ only by index.
            thread_key = conv_id or email.get("email_id", "")
            new_count  = 0

            for prod_idx, extracted in enumerate(product_records, start=1):
                if not any(v for v in extracted.values()):
                    continue  # skip genuinely empty records

                req_id = _make_req_id(thread_key, prod_idx)
                record = {
                    "req_id":                 req_id,
                    "status":                 "New",
                    "received_date":          email.get("received_date", "")[:10],
                    "sender_name":            sender_name,
                    "sender_email":           sender_email,
                    "email_subject":          email.get("subject", ""),
                    "assigned_to":            assigned_to,
                    "assigned_to_confidence": assigned_to_confidence,
                }
                record.update(extracted)

                brand    = record.get("product_brand") or record.get("product_category") or "(unknown)"
                customer = record.get("customer") or "(unknown)"
                qty      = record.get("quantity") or "?"
                unit     = record.get("quantity_unit") or ""
                print(f"           -> [{prod_idx}] {brand} | {customer} | qty={qty} {unit}")

                records.append(record)
                new_count += 1

            if new_count == 0:
                print("           WARNING: all extracted records were empty — skipping.")
                failed += 1

        except Exception as e:
            failed += 1
            print(f"           ERROR: {type(e).__name__}: {e} — logged, continuing")

        print()
        # Groq free tier: ~6000 tokens/min. Each thread email can be 1500-2500 tokens.
        # 15s between emails keeps total usage well under the per-minute token limit.
        if i < len(zavenir_emails):
            time.sleep(15)

    if failed:
        print(f"  {failed} email(s) failed extraction — see output above.")

    return {
        "records":        records,
        "emails_fetched": len(emails),
        "zavenir_found":  len(zavenir_emails),
    }


# ===========================================================================
# SENDER FILTER HELPER
# ===========================================================================

def _is_zavenir_email(email: dict) -> bool:
    """
    Return True if this email originated from tarora@zavenir.com.

    Two checks, in order:
      1. sender_email field matches directly — catches emails sent by Tania
         directly to the monitored inbox.
      2. body_content contains the address — catches emails where Saral
         forwarded Tania's email; the original From: header appears in the
         body text even though sender_email shows Saral's address.
    """
    # Check 1: direct sender match
    if email.get("sender_email", "").lower() == SENDER_FILTER.lower():
        return True

    # Check 2: forwarded-email body scan
    body = email.get("body_content", "")
    return bool(_ZAVENIR_EMAIL_RE.search(body))


# ===========================================================================
# REQ_ID GENERATION
# ===========================================================================

def _make_req_id(email_id: str, idx: int = 1) -> str:
    """
    Generate a unique record ID in the same format as the GET Global pipeline.

    Format: REQ-YYYYMMDD-XXXXXX-NN
      YYYYMMDD — today's date
      XXXXXX   — first 6 hex chars of MD5 hash of email_id / conversationId (uppercase)
      NN       — product index within the thread (01, 02, 03 …)

    Using conversationId as the hash base ensures all products from the same
    thread share a common prefix; the index makes each product unique.
    MD5 is deterministic — re-running the same thread always produces the
    same req_ids, so Supabase upsert is idempotent.
    """
    id_hash  = hashlib.md5(email_id.encode()).hexdigest()[:6].upper() if email_id else "000000"
    date_str = datetime.date.today().strftime("%Y%m%d")
    return f"REQ-{date_str}-{id_hash}-{idx:02d}"


# ===========================================================================
# SAVE — EXCEL
# ===========================================================================

def _save_to_excel(records: list) -> int:
    """
    Save Zavenir records to data/zavenir_crm.xlsx.

    - Creates the file (and data/ folder) if they do not exist yet.
    - Uses the FIELDS list from configs/zavenir_daubert.py as the column schema.
    - Deduplicates on req_id — skips any record whose req_id already exists.
    - Returns the number of new rows actually written.
    """
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    # Build a DataFrame from the records list, filling any missing schema
    # columns with None so the Excel file always has all columns in the right order.
    new_df = pd.DataFrame(records)
    for col in FIELDS:
        if col not in new_df.columns:
            new_df[col] = None
    new_df = new_df[FIELDS]   # enforce column order from schema

    if EXCEL_PATH.exists():
        existing_df  = pd.read_excel(EXCEL_PATH, dtype=str).fillna("")
        existing_ids = set(existing_df["req_id"].astype(str))
        truly_new    = new_df[~new_df["req_id"].astype(str).isin(existing_ids)]
        combined_df  = pd.concat([existing_df, truly_new], ignore_index=True)
        saved_count  = len(truly_new)
    else:
        combined_df = new_df
        saved_count = len(new_df)

    combined_df.to_excel(EXCEL_PATH, index=False)
    _format_excel()
    print(f"Excel: {saved_count} new record(s) saved -> {EXCEL_PATH.name}")
    return saved_count


def _format_excel():
    """Apply dark-header formatting to zavenir_crm.xlsx (matches GET Global style)."""
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)

    ws.freeze_panes = "A2"

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    wb.save(EXCEL_PATH)


# ===========================================================================
# SAVE — SUPABASE
# ===========================================================================

def _save_to_supabase(records: list) -> int:
    """
    Upsert Zavenir records into the zavenir_requirements Supabase table.

    Uses req_id as the conflict key — running the pipeline twice on the same
    email will update the existing row rather than creating a duplicate.
    Returns the number of records successfully upserted.
    """
    if not supabase:
        print("  Supabase not connected — cloud save skipped.")
        return 0

    print(f"\nPushing {len(records)} record(s) to Supabase ({SUPABASE_TABLE})...")
    saved = 0

    for record in records:
        # Build the Supabase row — keys must exactly match the table column names
        # in scripts/create_zavenir_table.sql.
        sb_row = {
            "req_id":           record.get("req_id"),
            "customer":         record.get("customer"),
            "industry_segment": record.get("industry_segment"),
            "product_category": record.get("product_category"),
            "product_brand":    record.get("product_brand"),
            "quantity":         record.get("quantity"),
            "quantity_unit":    record.get("quantity_unit"),
            "location":         record.get("location"),
            "delivery_date":    record.get("delivery_date"),
            "status":           record.get("status", "New"),
            "received_date":    record.get("received_date"),
            "sender_name":      record.get("sender_name"),
            "sender_email":     record.get("sender_email"),
            "email_subject":    record.get("email_subject"),
            "email_summary":    record.get("email_summary"),
            "next_action":      record.get("next_action"),
            "llm_confidence":   _safe_float(record.get("llm_confidence")),
            "assigned_to":              record.get("assigned_to"),
            "assigned_to_confidence":   record.get("assigned_to_confidence"),
            "conversation_timeline":    record.get("conversation_timeline"),
        }

        # Drop None values so Supabase does not overwrite existing data with NULL
        # on a re-run where only some fields changed.
        sb_row = {k: v for k, v in sb_row.items() if v is not None}

        try:
            supabase.table(SUPABASE_TABLE).upsert(
                sb_row, on_conflict="req_id"
            ).execute()
            print(f"   Supabase saved: {record.get('email_subject', 'no subject')[:55]}")
            saved += 1
        except Exception as e:
            print(f"   Supabase save failed for {record.get('req_id', '?')}: {e}")

    print("Supabase sync complete.")
    return saved


def save_records(records: list) -> dict:
    """
    Save all records to Excel and Supabase.
    Returns dict with excel_saved and supabase_saved counts.
    """
    if not records:
        print("  No records to save.")
        return {"excel_saved": 0, "supabase_saved": 0}

    excel_saved    = _save_to_excel(records)
    supabase_saved = _save_to_supabase(records)

    return {"excel_saved": excel_saved, "supabase_saved": supabase_saved}


# ===========================================================================
# UTILITY
# ===========================================================================

def _safe_float(value):
    """Convert llm_confidence to float, return None on failure."""
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


# ===========================================================================
# ENTRY POINT
# ===========================================================================

if __name__ == "__main__":

    pipeline_result = run_pipeline(max_emails=50)

    records        = pipeline_result["records"]
    emails_fetched = pipeline_result["emails_fetched"]
    zavenir_found  = pipeline_result["zavenir_found"]

    if DRY_RUN:
        print()
        print("=" * 65)
        print("  DRY RUN SUMMARY (no saves)")
        print("=" * 65)
        print(f"  Emails fetched          : {emails_fetched}")
        print(f"  Zavenir emails found    : {zavenir_found}")
        print(f"  Records extracted       : {len(records)}")
        print(f"  (Set DRY_RUN = False to save to Excel + Supabase)")
    else:
        print(f"[Step 3] Saving {len(records)} record(s) to Excel + Supabase...")
        save_result = save_records(records)

        print()
        print("=" * 65)
        print("  ZAVENIR RUN SUMMARY")
        print("=" * 65)
        print(f"  Emails fetched          : {emails_fetched}")
        print(f"  Zavenir emails found    : {zavenir_found}")
        print(f"  Records extracted       : {len(records)}")
        print(f"  Saved to Excel          : {save_result['excel_saved']}")
        print(f"  Saved to Supabase       : {save_result['supabase_saved']}")
        if EXCEL_PATH.exists():
            print(f"\n  Open: {EXCEL_OUTPUT}")
